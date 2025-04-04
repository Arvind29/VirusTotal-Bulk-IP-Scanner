import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

API_KEY = "ba711c14d8cec5b2e184d59074ac8849a1a1a075656b3d9c689acde807f1b43d"
VT_URL = "https://www.virustotal.com/api/v3/ip_addresses/"

cwd = os.getcwd()
file_name = "input.xlsx"
file_path = os.path.join(cwd, file_name)

try:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_name} not found in {cwd}")

    df = pd.read_excel(file_path)
    if df.empty or "IP" not in df.columns:
        raise ValueError("Excel must have 'IP' column.")

    def convert_date(timestamp):
        return datetime.utcfromtimestamp(int(timestamp)).strftime('%m/%d/%Y') if timestamp else ""

    def get_vt_data(ip):
        try:
            headers = {"x-apikey": API_KEY}
            response = requests.get(VT_URL + ip, headers=headers)
            if response.status_code != 200:
                return None

            data = response.json().get("data", {}).get("attributes", {})
            return {
                "Country": data.get("country", ""),
                "AS Owner": data.get("as_owner", ""),
                "RIR": data.get("regional_internet_registry", ""),
                "WHOIS Date": convert_date(data.get("whois_date")),
                "Last Analysis Date": convert_date(data.get("last_analysis_date")),
                "Malicious": data.get("last_analysis_stats", {}).get("malicious", 0),
                "Suspicious": data.get("last_analysis_stats", {}).get("suspicious", 0),
                "Undetected": data.get("last_analysis_stats", {}).get("undetected", 0),
                "Harmless": data.get("last_analysis_stats", {}).get("harmless", 0),
                "Timeout": data.get("last_analysis_stats", {}).get("timeout", 0),
            }
        except:
            return None

    output_data = []
    for ip in df["IP"]:
        if pd.notna(ip):
            result = get_vt_data(str(ip))
            if result:
                output_data.append([ip] + list(result.values()))

    output_file = "output.xlsx"
    file_path = os.path.join(cwd, output_file)

    columns = ["IP", "Country", "AS Owner", "RIR", "WHOIS Date", "Last Analysis Date",
               "Malicious", "Suspicious", "Undetected", "Harmless", "Timeout"]
    pd.DataFrame(output_data, columns=columns).to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active
    threshold = 2
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            if cell.value and isinstance(cell.value, int) and cell.value > threshold:
                ws.cell(row=cell.row, column=1).fill = red_fill

    wb.save(output_file)
    print(f"Process completed. Results saved in {output_file}")

except Exception as e:
    print(f"Error: {e}")
